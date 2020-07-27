import os
from os import path
import csv
from pprint import pprint
import rpa as r
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import mimetypes 

# путь к текущей папке скрипта
scr_dir = path.dirname(path.abspath(__file__))
# пути к csv файлам
csv_file_usd = scr_dir + r'\table_usd.csv'
csv_file_eu = scr_dir + r'\table_eu.csv'
# пути к xlsx файлам
xlsx_file_out = scr_dir + r'\table.xlsx'

r.init(visual_automation = True)

r.url('https://yandex.ru/news/quotes/2002.html')
r.table('quote__data', csv_file_usd)

r.url('https://yandex.ru/news/quotes/2000.html')
r.table('quote__data', csv_file_eu)

r.close()

workbook = Workbook()
sheet = workbook.active

# финансовый рублёвый формат
cell_format = u'# ##0.00 ₽;[RED]-# ##0.00 ₽'

# числовой формат
num_format = u'# ##0.00;[RED]-# ##0.00'

# функция для суммы значений столбцов
def summ_cells(row, col):
    final_row = row 
    cell = sheet.cell(row, col)
    while cell.value is not None:
        final_row += 1
        cell = sheet.cell(final_row, col)
    start = get_column_letter(col) + str(row)
    finish = get_column_letter(col) + str(final_row-1)
    cell.value = f'=SUM({start}:{finish})'
    cell.number_format = num_format
    return final_row

# функция для переноса csv в файл excel
def addCsvToXlsx(csv_file, cols):
    # читаем csv файл usd
    with open(csv_file, 'r', encoding='utf-8') as file:
        reader = csv.reader(file, delimiter=',')
        row = 1
        for line in reader:
            c = 0
            for col in cols:
                cell = sheet.cell(row, col)
                # если не столбец даты, то применяем формат
                if row > 1 and c > 0:
                    cell.value = float(line[c].replace(',', '.'))
                    cell.number_format = cell_format
                else:
                    cell.value = line[c]
                c += 1
            row += 1

# читаем csv файлы и добавляем в итоговый xlsx
addCsvToXlsx(csv_file_usd, [1, 2, 3])
addCsvToXlsx(csv_file_eu, [4, 5, 6])

# создаём столбец для разности
sheet['G1'] = 'Динамика'
for row in range(2, 12):
    cell = sheet.cell(row, 7)
    strrow = str(row)
    cell.value = '=E' + strrow + '/B' + strrow
    cell.number_format = num_format

# ставим автоматические размеры столбцам
for col in range(1,8):
    sheet.column_dimensions[get_column_letter(col)].auto_size = True


summ_cells(2, 3)
last_row = summ_cells(2, 6)

workbook.save(filename=xlsx_file_out)

os.system("start " + xlsx_file_out)

# Отправить на почту сообщение с файлом
addr_from = ""
addr_to   = ""
password  = ""

msg = MIMEMultipart()
msg['From']    = addr_from
msg['To']      = addr_to
msg['Subject'] = 'Курс USD и EU за 10 дней'

body = f"Количество строк: {last_row}"
msg.attach(MIMEText(body, 'plain'))

ctype, encoding = mimetypes.guess_type(xlsx_file_out)
if ctype is None or encoding is not None:
    ctype = 'application/octet-stream'
maintype, subtype = ctype.split('/', 1)


with open(xlsx_file_out, 'rb') as fp:
    file = MIMEBase(maintype, subtype)
    file.set_payload(fp.read())
    fp.close()
encoders.encode_base64(file)

file.add_header('Content-Disposition', 'attachment', filename=os.path.basename(xlsx_file_out))
msg.attach(file)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(addr_from, password)
server.send_message(msg)
server.quit()